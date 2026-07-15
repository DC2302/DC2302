"""Write the outputs: leads.xlsx (Excel / Google Sheets), leads.csv,
and call_sheet.md (a readable who-to-call list with talking points)."""

import csv
from datetime import date
from pathlib import Path

CONTACT_HELP = """\
## How to get a phone number for these companies

1. **Texas operators:** look the company up in the RRC's organization
   (P-5) query -- it lists the registered address and phone number for
   every active operator in Texas:
   https://webapps2.rrc.texas.gov/EWA/organizationQueryAction.do
2. **New Mexico operators:** search the OGRID in OCD's operator lookup:
   https://wwwapps.emnrd.nm.gov/ocd/ocdpermitting/Data/Operators.aspx
3. LinkedIn: search the company name plus "facilities engineer",
   "construction superintendent", or "completions/production
   superintendent" -- those are the people who hire facility
   constructors like Lobos.
4. Ask for: the **facilities engineering** or **construction** group,
   not the drilling department.
"""


def write_leads_csv(leads: list, path: Path):
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            [
                "rank", "operator", "category", "score", "drilling_permits",
                "swd_permits", "pipeline_permits", "datacenter_projects",
                "counties", "states", "latest_activity", "pitch",
            ]
        )
        for i, lead in enumerate(leads, 1):
            writer.writerow(
                [
                    i,
                    lead["operator"],
                    lead["category"],
                    lead["score"],
                    lead["drilling_permits"],
                    lead["swd_permits"],
                    lead["pipeline_permits"],
                    lead["datacenter_projects"],
                    "; ".join(lead["counties"]),
                    "; ".join(lead["states"]),
                    lead["latest_activity"],
                    lead["pitch"],
                ]
            )


def write_permits_csv(permits: list, path: Path):
    if not permits:
        return
    cols = ["signal", "state", "operator", "county", "district",
            "date", "purpose", "lease", "permit_no",
            "contractor", "design_firm", "cost"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(permits)


LEAD_COLUMNS = [
    ("Rank", 6), ("Company", 38), ("Category", 12), ("Score", 7),
    ("Drilling permits", 15), ("SWD permits", 12),
    ("Pipeline permits", 14), ("Data center projects", 12),
    ("Counties", 28), ("States", 8), ("Latest activity", 14),
    ("What to pitch", 90),
    ("Phone", 16), ("Contact name", 22), ("Call notes / status", 40),
]


def write_leads_xlsx(leads: list, permits: list, path: Path):
    """Excel workbook: a formatted 'Leads' tab (with empty phone/notes
    columns to fill in as you call) plus a 'Raw permits' tab.
    Opens directly in Excel and imports cleanly into Google Sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, (title, width) in enumerate(LEAD_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = width

    swd_fill = PatternFill("solid", fgColor="FFF2CC")   # SWD leads
    dc_fill = PatternFill("solid", fgColor="DDEBF7")    # data-center leads
    pitch_col = next(
        i for i, (title, _) in enumerate(LEAD_COLUMNS, 1)
        if title == "What to pitch"
    )
    for i, lead in enumerate(leads, 2):
        row = [
            i - 1,
            lead["operator"],
            "Data center" if lead["category"] == "datacenter" else "Oil & gas",
            lead["score"],
            lead["drilling_permits"],
            lead["swd_permits"],
            lead["pipeline_permits"],
            lead["datacenter_projects"],
            ", ".join(lead["counties"]),
            "/".join(lead["states"]),
            lead["latest_activity"],
            lead["pitch"],
            "", "", "",  # phone / contact / notes -- fill in while calling
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.alignment = Alignment(
                vertical="top", wrap_text=(col == pitch_col)
            )
        fill = None
        if lead["category"] == "datacenter":
            fill = dc_fill
        elif lead["swd_permits"]:
            fill = swd_fill
        if fill:
            for col in range(1, len(row) + 1):
                ws.cell(row=i, column=col).fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(LEAD_COLUMNS))}{len(leads) + 1}"

    raw = wb.create_sheet("Raw permits")
    raw_cols = ["signal", "state", "operator", "county", "district",
                "date", "purpose", "lease", "permit_no",
                "contractor", "design_firm", "cost"]
    for col, title in enumerate(raw_cols, 1):
        cell = raw.cell(row=1, column=col, value=title.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        raw.column_dimensions[get_column_letter(col)].width = 22
    for r, permit in enumerate(permits, 2):
        for col, key in enumerate(raw_cols, 1):
            raw.cell(row=r, column=col, value=permit.get(key, ""))
    raw.freeze_panes = "A2"

    wb.save(path)


def write_call_sheet(leads: list, path: Path, top_n: int = 25,
                     demo: bool = False):
    lines = [
        f"# Lobos lead call sheet -- {date.today():%B %d, %Y}",
        "",
    ]
    if demo:
        lines += [
            "> **DEMO DATA** -- this sheet was generated from bundled",
            "> sample data so you can see the format. Run without",
            "> `--demo` to pull live permit data.",
            "",
        ]
    oilgas = [l for l in leads if l["category"] == "oilgas"]
    datacenter = [l for l in leads if l["category"] == "datacenter"]

    lines += [
        f"{len(leads)} companies with recent activity, ranked by how "
        "much facility-construction work their filings imply.",
        "",
        "# Part 1 -- Oil & gas operators (Permian Basin)",
        "",
    ]
    lines += _lead_entries(oilgas[:top_n])

    if datacenter:
        lines += [
            "# Part 2 -- Data center / large commercial builds "
            "(West Texas)",
            "",
            "These are construction projects, not operators: call the "
            "owner's development team AND the general contractor / "
            "design firm about subcontract packages (site work, "
            "trenching, utility infrastructure, concrete).",
            "",
        ]
        lines += _lead_entries(datacenter[:top_n])

    lines += [CONTACT_HELP]
    path.write_text("\n".join(lines), encoding="utf-8")


def _lead_entries(leads: list) -> list:
    lines = []
    for i, lead in enumerate(leads, 1):
        activity = []
        if lead["drilling_permits"]:
            activity.append(f"{lead['drilling_permits']} drilling permit(s)")
        if lead["swd_permits"]:
            activity.append(f"{lead['swd_permits']} SWD permit(s)")
        if lead["pipeline_permits"]:
            activity.append(f"{lead['pipeline_permits']} pipeline permit(s)")
        if lead["datacenter_projects"]:
            activity.append(
                f"{lead['datacenter_projects']} registered construction "
                "project(s)"
            )
        lines += [
            f"## {i}. {lead['operator']}  (score {lead['score']})",
            "",
            f"- **Activity:** {', '.join(activity) or 'n/a'}",
            f"- **Where:** {', '.join(lead['counties'])} "
            f"({'/'.join(lead['states'])})",
            f"- **Most recent:** {lead['latest_activity'] or 'n/a'}",
            f"- **Talking point:** {lead['pitch']}",
            "",
        ]
    return lines
